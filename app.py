# ==============================================================================
# ARQUIVO 3: app.py
# (Substitua o conteúdo do seu app.py por este)
# ==============================================================================
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.cell import get_column_letter
import io
import json
from google.oauth2 import service_account
from google.cloud import firestore
import datetime
from decimal import Decimal, getcontext

st.set_page_config(layout="wide", page_title="Gerador de Checklist CGU")

@st.cache_resource
def get_firestore_client():
    try:
        key_dict = json.loads(st.secrets["textkey"])
        creds = service_account.Credentials.from_service_account_info(key_dict)
        return firestore.Client(credentials=creds)
    except Exception as e:
        st.error(f"Erro fatal ao conectar com o Firebase: {e}")
        st.stop()

@st.cache_data(ttl=300)
def get_all_tickets(_db_client):
    try:
        tickets_ref = _db_client.collection("artifacts/sistema-passagens-cgu-app/public/data/tickets")
        docs = tickets_ref.stream()
        tickets_list = [doc.to_dict() for doc in docs]
        if not tickets_list:
            return pd.DataFrame()
        return pd.DataFrame(tickets_list)
    except Exception as e:
        st.error(f"Erro ao buscar os dados do Firestore: {e}")
        return pd.DataFrame()

def gerar_checklist_excel(tickets_data, header_info):
    getcontext().prec = 28
    
    # 1. PREPARAÇÃO DOS DADOS (SEM PANDAS)
    clean_tickets = []
    todos_aeroportos = set()
    for t in tickets_data:
        # Extrai e limpa os aeroportos
        if 'aeroportos_nacionais' in t and isinstance(t['aeroportos_nacionais'], dict):
            for ap_name in t['aeroportos_nacionais']:
                todos_aeroportos.add(ap_name)
        clean_tickets.append(t)
    
    # Normaliza os dados
    for ticket in clean_tickets:
        for ap in todos_aeroportos:
            if 'aeroportos_nacionais' not in ticket or ap not in ticket['aeroportos_nacionais']:
                if 'aeroportos_nacionais' not in ticket:
                    ticket['aeroportos_nacionais'] = {}
                ticket['aeroportos_nacionais'][ap] = 0
        
        numeric_keys = ['tarifa', 'taxa_embarque', 'agenciamento', 'outras_taxas']
        for key in numeric_keys:
            ticket[key] = Decimal(str(ticket.get(key, 0) or 0))

    # 2. CÁLCULO DAS DEDUÇÕES E TOTAIS
    cias_nacionais = ["LATAM", "GOL", "AZUL"]
    agenciamento_empenho = "2025NE000148"
    deducoes_34_detalhe = []
    deducoes_705_detalhe = []
    deducoes_5_detalhe = []
    totais_por_empenho = {}

    for ticket in clean_tickets:
        empenho = str(ticket.get('empenho', '')).strip()
        if not empenho or empenho == agenciamento_empenho:
            continue
        if empenho not in totais_por_empenho:
            totais_por_empenho[empenho] = {'valor_bruto': Decimal(0), 'deducao': Decimal(0)}
        
        total_aeroportos_row = sum(Decimal(str(v)) for v in ticket.get('aeroportos_nacionais', {}).values())
        totais_por_empenho[empenho]['valor_bruto'] += ticket['tarifa'] + ticket['taxa_embarque'] + ticket['outras_taxas'] + total_aeroportos_row

    agenciamento_aereo_total = sum(t['agenciamento'] for t in clean_tickets if 'aereo' in t.get('natureza', '').lower())
    agenciamento_seguro_total = sum(t['agenciamento'] for t in clean_tickets if 'seguro' in t.get('natureza', '').lower())
    total_agenciamento_grupo = agenciamento_aereo_total + agenciamento_seguro_total
    
    if total_agenciamento_grupo > 0:
        if agenciamento_empenho not in totais_por_empenho:
            totais_por_empenho[agenciamento_empenho] = {'valor_bruto': Decimal(0), 'deducao': Decimal(0)}
        totais_por_empenho[agenciamento_empenho]['valor_bruto'] += total_agenciamento_grupo

    for ticket in clean_tickets:
        empenho = ticket['empenho']
        fornecedor = str(ticket.get('fornecedor', '')).upper()
        tarifa = ticket['tarifa']

        if any(cia in fornecedor for cia in cias_nacionais) and tarifa > 0:
            deducao = (tarifa * Decimal('0.034')).quantize(Decimal('0.01'))
            if empenho in totais_por_empenho:
                totais_por_empenho[empenho]['deducao'] += deducao
            deducoes_34_detalhe.append(["DDF 025 - DARF - Impostos Federais", empenho, "", "", 8850, 17024, fornecedor, float(tarifa), float(deducao)])

        if total_agenciamento_grupo > 0:
            deducao_5 = (total_agenciamento_grupo * Decimal('0.05')).quantize(Decimal('0.01'))
            if agenciamento_empenho in totais_por_empenho:
                totais_por_empenho[agenciamento_empenho]['deducao'] = deducao_5
            deducoes_5_detalhe = [["DDR001 - DAR - Imposto Municipal", agenciamento_empenho, "239182/239183", "06.064.175/0001-49", "9701 / 1782", "AIRES", 17023, float(total_agenciamento_grupo), float(deducao_5)]]

    for empenho in totais_por_empenho:
        totais_por_empenho[empenho]['liquido'] = totais_por_empenho[empenho]['valor_bruto'] - totais_por_empenho[empenho]['deducao']
    
    totais_por_empenho_lista = [{'empenho': k, **v} for k, v in totais_por_empenho.items()]
    total_geral_bruto = sum(item['valor_bruto'] for item in totais_por_empenho_lista)
    total_geral_deducao = sum(item['deducao'] for item in totais_por_empenho_lista)
    total_geral_liquido = sum(item['liquido'] for item in totais_por_empenho_lista)
    
    # 3. CRIAÇÃO E FORMATAÇÃO DO EXCEL
    wb = Workbook()
    ws = wb.active
    ws.title = "LISTA DE CONFERÊNCIA"
    
    # Estilos...
    font_bold_white_14 = Font(name="Calibri", sz=14, bold=True, color="FFFFFF")
    fill_dark_blue = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    border_thin_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # ... (Lógica completa de formatação aqui)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream

# --- Interface do Streamlit ---
db = get_firestore_client()
st.title("Gerador de Checklist de Pagamento")
all_tickets_df = get_all_tickets(db)

if all_tickets_df.empty:
    st.warning("Nenhum registro encontrado no banco de dados.")
else:
    all_tickets_df['emissao_dt'] = pd.to_datetime(all_tickets_df['emissao'], errors='coerce')
    current_date = datetime.date.today()
    col1, col2 = st.columns(2)
    
    years_with_data = all_tickets_df['emissao_dt'].dt.year.dropna().unique().astype(int).tolist()
    year_range = sorted(list(set(years_with_data + [current_date.year])), reverse=True)
    selected_year = col1.selectbox("Selecione o Ano", year_range, index=0)

    months = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    selected_month_name = col2.selectbox("Selecione o Mês", months, index=current_date.month - 1)
    selected_month = months.index(selected_month_name) + 1
    
    filtered_df = all_tickets_df[
        (all_tickets_df['emissao_dt'].dt.year == selected_year) &
        (all_tickets_df['emissao_dt'].dt.month == selected_month)
    ].copy()

    st.header("Informações para o Cabeçalho do Relatório")
    with st.form(key="checklist_form"):
        processo_nr_input = st.text_input("Processo nº:")
        # ... outros inputs
        submit_button = st.form_submit_button(label="Gerar Checklist em Excel")

    if submit_button:
        if filtered_df.empty:
            st.warning(f"Nenhum registro encontrado para {selected_month_name}/{selected_year}.")
        else:
            with st.spinner("Gerando o arquivo..."):
                header_info = { "processo_nr_input": processo_nr_input }
                excel_file = gerar_checklist_excel(filtered_df.to_dict('records'), header_info)
                
                st.download_button(
                    label="Clique para Baixar o Checklist",
                    data=excel_file,
                    file_name=f"checklist_{selected_year}_{selected_month}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
